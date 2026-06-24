# -*- coding: utf-8 -*-
"""acceptance.md (engineering-version Gherkin) -> planner-version acceptance checklist xlsx, with optional real-value substitution.

Automatically translates `-05acceptance.md` (scenarios tagged with R-codes, Given/When/Then phrasing) into a checklist
that planners/QA "tick item by item while playing the game": an info page + a test-checklist page (dropdown pass/fail/pending,
frozen header, filter, progress-statistics formulas). **Cases are extracted automatically from Gherkin, not hand-listed.**

Real-value substitution (--config): a `Table[primary key].field` (e.g. `levelTable[50].value`) in an assertion is
looked up in the config tables and replaced with the real value (`500`), recording the source in the "source field" column. This way
planners get real numbers to verify against the game directly, while at the same time **reverse-validating that the engineering
version / rules read the right value from the table** (config<->rule consistency).

Guardrails (never fabricate):
  - single primary key and numeric (`levelTable[50]`) -> look up col0/id column and substitute the real value.
  - multi primary key / enum-name key (`starTable[B,fire,0]`) -> mark `[manual fill]`, **do not guess**.
  - not found (table/field/row missing) -> mark `[not found]`, keep the original assertion text.
  - when --loc gives LocalizationText, text fields like NameId/DescId additionally get the Chinese appended.

Usage:
  python gherkin_to_checklist.py <acceptance.md> [out.xlsx] [--config <config dir>] [--loc <LocalizationText.xlsx>]
Depends on openpyxl (write side); with --config, reuses the same-directory xlsx_dump.py to parse config tables.
"""
import sys, os, re, datetime, json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import xlsx_dump   # reuse zipfile+xml parsing (openpyxl errors when reading domestic xlsx)
import zipfile
from config_index import build_index, load_enums, load_keymap, lookup, REF_RE   # shared index/resolution layer

FONT = "Microsoft YaHei"
HEAD_FILL  = PatternFill("solid", fgColor="2F7D18")
GRP_FILL   = PatternFill("solid", fgColor="EAF4E0")
TITLE_FILL = PatternFill("solid", fgColor="1D3A17")
thin = Side(style="thin", color="C9D6BD")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

GIVEN = ("Given",)
WHEN  = ("When",)
THEN  = ("Then",)
CONT  = ("And", "But")

# ---------- Gherkin parsing ----------
def extract_r(text):
    m = re.search(r"[（(]\s*(R-[A-Za-z0-9\-/ ]+?)\s*[)）]\s*$", text)
    return (text[:m.start()].strip(), m.group(1).strip()) if m else (text.strip(), "")

def parse(md_text):
    feature, cases, cur, last = "", [], None, None
    def flush():
        if cur and (cur["given"] or cur["when"] or cur["then"] or cur["name"]):
            cases.append(cur)
    for raw in md_text.splitlines():
        line = raw.strip()
        if not line: continue
        m = re.match(r"^#*\s*Feature\s*[:：]\s*(.+)$", line)
        if m: feature = extract_r(m.group(1))[0]; continue
        m = re.match(r"^#*\s*Scenario\s*[:：]\s*(.+)$", line)
        if m:
            flush(); name, r = extract_r(m.group(1))
            cur = {"feature": feature, "name": name, "r": r, "given": [], "when": [], "then": []}
            last = cur["then"]; continue
        if cur is None: continue
        hit = None
        for kw in GIVEN:
            if line.startswith(kw): hit = ("given", line[len(kw):].strip()); break
        if not hit:
            for kw in WHEN:
                if line.startswith(kw): hit = ("when", line[len(kw):].strip()); break
        if not hit:
            for kw in THEN:
                if line.startswith(kw): hit = ("then", line[len(kw):].strip()); break
        if not hit:
            for kw in CONT:
                if line.startswith(kw):
                    if last is not None: last.append(line[len(kw):].strip())
                    hit = ("cont", None); break
        if hit and hit[0] != "cont":
            cur[hit[0]].append(hit[1]); last = cur[hit[0]]
    flush(); return cases

def title_of(md_text, fallback):
    for line in md_text.splitlines():
        m = re.match(r"^#\s+(.+)$", line.strip())
        if m: return re.split(r"[·\-—|（(]", m.group(1).strip())[0].strip()
    return fallback

# ---------- config-table index + value lookup (build_index / load_enums / load_keymap / lookup, see config_index) ----------

def is_textfield(field):
    f = field.lower()
    return ("name" in f) or ("desc" in f)

def subst(text, idx, loc):
    """Substitute real values, return (new text, [source entries]). idx=None means only extract references, do not substitute."""
    sources = []
    def repl(m):
        ref, table, keystr, field = m.group(0), m.group(1), m.group(2).strip(), m.group(3)
        if idx is None:
            sources.append(ref); return ref
        val = lookup(idx, table, keystr, field)
        if val is None:
            sources.append(f"{ref} [not found]"); return ref
        if val == "MULTI":
            sources.append(f"{ref} [multi-key/enum, manual fill]"); return ref
        cn = loc.get(val, "") if (loc and re.fullmatch(r"\d+", val) and is_textfield(field)) else ""
        shown = val + (f"（{cn}）" if cn else "")
        sources.append(f"{ref}={shown}")
        return shown
    return REF_RE.sub(repl, text), sources

# ---------- xlsx generation ----------
def style_header(c):
    c.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
    c.fill = HEAD_FILL
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BORDER

def build(cases, title, out, idx, loc):
    join = lambda xs: "；".join(xs) if xs else "—"
    wb = Workbook()
    s = wb.active; s.title = "Info"; s.sheet_view.showGridLines = False
    s.merge_cells("A1:E1")
    s["A1"] = f"{title} . Acceptance test checklist (planner-readable version)"
    s["A1"].font = Font(name=FONT, bold=True, color="FFFFFF", size=14)
    s["A1"].fill = TITLE_FILL
    s["A1"].alignment = Alignment(horizontal="center", vertical="center")
    s.row_dimensions[1].height = 30
    today = datetime.date.today().isoformat()
    note = (f"{today} auto-generated by gherkin_to_checklist.py from -05acceptance.md, {len(cases)} cases in total"
            + ("(config real values substituted, see the 'source field' column; [manual fill]/[not found] are guardrail marks)." if idx else "(no config attached, assertions keep engineering field references)."))
    rows = [
        ("", ""),
        ("How to use", "(1) set the game to the specified state per 'precondition' -> (2) tap as per 'operation' -> (3) verify actual behavior against 'expected result', tick the actual-result/pass columns on the 'Test checklist' page."),
        ("Real value/source", "Numbers in 'expected' are substituted with config real values; the 'source field' column marks the source (`Table[primary key].field`). When config changes, just refresh from the source, the logic does not change. Can also be checked in reverse: does the value the engineering/rules read from the table match this."),
        ("Linkage", "Each case is tagged with an engineering-version R-code, linkable back to -05acceptance.md and -01system-rules.md."),
        ("Generation", note),
    ]
    r = 3
    for k, v in rows:
        s[f"A{r}"] = k; s[f"A{r}"].font = Font(name=FONT, bold=True, size=10, color="2F7D18")
        s.merge_cells(f"B{r}:E{r}")
        s[f"B{r}"] = v; s[f"B{r}"].font = Font(name=FONT, size=10)
        s[f"B{r}"].alignment = Alignment(wrap_text=True, vertical="center")
        s.row_dimensions[r].height = 38; r += 1
    r += 1
    s[f"A{r}"] = "Progress stats"; s[f"A{r}"].font = Font(name=FONT, bold=True, size=11, color="1D3A17"); r += 1
    last = len(cases) + 1
    stat = [("Total cases", f'=COUNTA(Test checklist!A2:A{last})'),
            ("Passed",   f'=COUNTIF(Test checklist!J2:J{last},"pass")'),
            ("Failed",   f'=COUNTIF(Test checklist!J2:J{last},"fail")'),
            ("Pending",     "=B{0}-B{1}-B{2}"),
            ("Pass rate",   "=IF(B{0}=0,0,B{1}/B{0})")]
    base = r
    for i, (k, f) in enumerate(stat):
        rr = base + i
        s[f"A{rr}"] = k; s[f"A{rr}"].font = Font(name=FONT, bold=True, size=10); s[f"A{rr}"].border = BORDER
        if "{0}" in f: f = f.format(base, base+1, base+2)
        s[f"B{rr}"] = f; s[f"B{rr}"].font = Font(name=FONT, size=10); s[f"B{rr}"].border = BORDER
        s[f"B{rr}"].alignment = Alignment(horizontal="center")
    s[f"B{base+4}"].number_format = "0.0%"
    for col, w in zip("ABCDE", [12, 22, 10, 10, 30]): s.column_dimensions[col].width = w

    t = wb.create_sheet("Test checklist"); t.sheet_view.showGridLines = False
    headers = ["No.", "Feature", "Test point", "Precondition", "Operation", "Expected result", "Source field", "Linked R-code", "Actual result", "Pass"]
    for i, h in enumerate(headers): style_header(t.cell(row=1, column=1+i, value=h))
    t.row_dimensions[1].height = 24
    for ri, c in enumerate(cases, start=2):
        g, sg = subst(join(c["given"]), idx, loc)
        w, sw = subst(join(c["when"]),  idx, loc)
        th, st = subst(join(c["then"]), idx, loc)
        src = []
        for x in sg + sw + st:
            if x not in src: src.append(x)
        row = [f"T{ri-1:02d}", c["feature"], c["name"], g, w, th, "；".join(src) or "—", c["r"]]
        for i, v in enumerate(row):
            cc = t.cell(row=ri, column=1+i, value=v)
            cc.font = Font(name=FONT, size=10); cc.border = BORDER
            cc.alignment = Alignment(vertical="center", wrap_text=True,
                                     horizontal="center" if i in (0, 7) else "left")
        t.cell(row=ri, column=2).fill = GRP_FILL
        for col in (9, 10):
            cc = t.cell(row=ri, column=col); cc.border = BORDER
            cc.alignment = Alignment(vertical="center", horizontal="center")
    for i, w in enumerate([7, 13, 19, 24, 14, 38, 26, 15, 18, 8]):
        t.column_dimensions[chr(65+i)].width = w
    t.freeze_panes = "A2"
    t.auto_filter.ref = f"A1:J{len(cases)+1}"
    dv = DataValidation(type="list", formula1='"pass,fail,pending"', allow_blank=True)
    t.add_data_validation(dv); dv.add(f"J2:J{len(cases)+1}")
    wb.save(out)

if __name__ == "__main__":
    args = sys.argv[1:]
    config_dir = loc_path = keymap_path = enums_path = None
    pos, i = [], 0
    while i < len(args):
        a = args[i]
        if   a == "--config" and i+1 < len(args): config_dir = args[i+1]; i += 2
        elif a == "--loc"    and i+1 < len(args): loc_path   = args[i+1]; i += 2
        elif a == "--keymap" and i+1 < len(args): keymap_path= args[i+1]; i += 2
        elif a == "--enums"  and i+1 < len(args): enums_path = args[i+1]; i += 2
        else: pos.append(a); i += 1
    if not pos:
        print("usage: python gherkin_to_checklist.py <acceptance.md> [out.xlsx] "
              "[--config <config dir>] [--enums <enum-dict.md>] [--keymap <composite-key-map.json>] [--loc <LocalizationText.xlsx>]"); sys.exit(1)
    src = pos[0]
    out = pos[1] if len(pos) > 1 else os.path.splitext(src)[0] + "-planner.xlsx"
    with open(src, "r", encoding="utf-8") as f: md = f.read()
    cases = parse(md)
    if not cases:
        print("WARN: no scenarios parsed (check the Gherkin 'Scenario:' format)"); sys.exit(2)
    idx = None
    if config_dir:
        idx = build_index(config_dir)
        if not keymap_path:                                   # composite-key map: --keymap takes priority, otherwise auto-find it in the config dir
            auto = os.path.join(config_dir, "composite-key-map.json")
            if os.path.exists(auto): keymap_path = auto
        if keymap_path:
            try: idx["__keymap__"] = load_keymap(keymap_path)
            except Exception as e: print("WARN: keymap parse failed:", e)
        if enums_path:
            try: idx["__enums__"] = load_enums(enums_path)
            except Exception as e: print("WARN: enums parse failed:", e)
    loc = None
    if loc_path:
        try:
            import resolve_loc
            loc = resolve_loc.build(loc_path)
        except Exception as e:
            print("WARN: --loc parse failed, skipping text substitution:", e)
    build(cases, title_of(md, os.path.basename(src)), out, idx, loc)
    extra = ""
    if idx is not None:
        ntab = sum(1 for k in idx if not k.startswith("__"))
        flags = sum(1 for c in cases for x in
                    subst("；".join(c["given"]+c["when"]+c["then"]), idx, loc)[1] if "[" in x)
        extra = (f" | config tables {ntab} | keymap {len(idx.get('__keymap__',{}))} tables"
                 f" | enums {len(idx.get('__enums__',{}))} names | guardrail marks {flags}")
    print(f"saved: {out} | cases: {len(cases)}{extra}")
